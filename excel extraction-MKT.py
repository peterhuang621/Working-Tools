from openpyxl import load_workbook
import os

filePath = r'C:\Users\ECI170\Desktop\qbr 模板\新\MKT\12'
summary_doc=load_workbook(r'C:\Users\ECI170\Desktop\qbr 模板\USD - copy MKT.xlsx')

def ext(x):
	doc=load_workbook(f'{filePath}\\{x}',data_only=True)
	def v(g):
		if doc[ln][g].value==None:
			return 0
		else:
			return doc[ln][g].value

	language_list=doc.sheetnames
	language_list.remove('All')
	for i in doc.sheetnames:
		if doc[i].sheet_state=='hidden':
			language_list.remove(i)
	print(x,':',language_list)
	for ln in language_list:
		a=summary_doc[ln]
		a[f'A{n}']=v('F13')
		a[f'C{n}']=v('D23')
		a[f'D{n}']=v('D24')
		a[f'E{n}']=v('D25')
		a[f'F{n}']=v('D26')
		a[f'G{n}']=v('D27')
		a[f'H{n}']=v('D28')
		a[f'I{n}']=v('D29')
		a[f'K{n}']=(v('D36')+v('D37')+v('D39'))*7+v('D38')*3
		a[f'L{n}']=v('D41')+v('D42')+v('D43')+v('D44')+v('D45')+v('D46')+v('D47')+v('D48')+v('D49')+v('D50')/60+v('D51')/60
		a[f'S{n}']=v('C57')
		a[f'T{n}']=v('C31')
		a[f'U{n}']=v('C34')+v('C55')
		a[f'V{n}']=v('C52')
		a[f'Y{n}']=v('F56')
	doc.close()

n=9
for i in os.listdir(f'{filePath}'):
	ext(i)
	n+=1

summary_doc.save(r'C:\Users\ECI170\Desktop\qbr 模板\USD MKT 12.xlsx')

