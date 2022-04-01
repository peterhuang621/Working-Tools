from openpyxl import load_workbook
import os

filePath = r'C:\Users\ECI170\Desktop\QBR Q4\LQA\MKT LQA\all'
##提取值的EXCEL位置

def ext(x):
	doc=load_workbook(f'{filePath}\\{x}',data_only=True)
	#用load_workbook加载每个文件，并声明只取值（而不是单元格的函数）
	def v(g):
		##定义一个sheet（比如我选的QA_results这个工作簿）取出值的函数方便后续操作，并且要求空值视作0
		if doc['QA_results'][g].value==None:
			return 0
		else:
			return doc['QA_results'][g].value

	total_num=0
	#统计用变量，初始值为0
	for i in range(2,201):
		#比如取的一列，列号从多少到多少
		cellx='E'+str(i)
		#取得列E，并且把列号数字加上
		if v(cellx)=='Preferential change':
			#关键：我设定取‘Preferential change’的值
			total_num+=1
			#每出现一次，统计变量加1
	print(doc['QA_form']['C3'].value,'|',total_num)
	#打印出每个LQA表的名字和出现v(cellx)的次数
	doc.close()


for i in os.listdir(f'{filePath}'):
	ext(i)
	#对每个EXCEL进行操作（提取）
