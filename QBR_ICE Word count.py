from openpyxl import load_workbook
import os

filePath = r'C:\Users\ECI170\Desktop\IVD(R)\Quotes\12\字数'
summary_doc=load_workbook(r'C:\Users\ECI170\Desktop\qbr 模板\Word_count.xlsx')
sn='Sheet1'

def ext(x):
	doc=load_workbook(f'{filePath}\\{x}',data_only=True)
	for i in doc.sheetnames:
		b=doc[i]
	a=summary_doc[sn]
	for r in range(2,10):
		for c in 'ABC':
			if b[f'{c}{r}'].value!=None:
				a[f'{c}{n+r}']=b[f'{c}{r}'].value
			else:
				a[f'{c}{n+r}']='None'
		a[f'D{n+r}']=x

	doc.close()

n=0
for i in os.listdir(f'{filePath}'):
	print(i)
	ext(i)
	n+=10

summary_doc.save(r'C:\Users\ECI170\Desktop\IVD(R)\Quotes\12\统计.xlsx')

